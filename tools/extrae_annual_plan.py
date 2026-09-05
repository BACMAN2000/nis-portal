# -*- coding: utf-8 -*-
"""Convierte el Annual Plan Primary 2026 del colegio en JSON.

El Excel vive en Drive (la carpeta NORDIC DOCUMENTS, dentro del ZIP "2. PRIMARY").
Cada hoja es un grado y la matriz es fila = semana, columna = area. Este
script NO interpreta nada: copia el texto tal cual esta escrito, y le pega
el numero de periodo y las semanas reales del Primary Calendar 2026.

1.o no esta en ese Excel: su matriz vive dentro de "EY_Assessment Criteria_G1.xlsx"
(hoja "Annual Plan"), porque 1.o pertenece a Early Years. Se pasa como segundo
argumento y sale en el mismo JSON.

Uso:  python tools/extrae_annual_plan.py <Annual Plan Primary.xlsx> [EY_Criteria_G1.xlsx] [salida.json]
"""
import json, re, sys, unicodedata
from pathlib import Path
import openpyxl

# Del Primary Calendar 2026. El Excel del plan anual a veces pone 6 filas
# donde el calendario solo da 5 semanas de clase (5.o lo anota a mano:
# "Solo son 5 semanitas"). Manda el calendario.
PERIODOS = [
    dict(p=1, unidad=1, trimestre=1, semanas=6, inicio='2026-03-09', fin='2026-04-17'),
    dict(p=2, unidad=2, trimestre=1, semanas=5, inicio='2026-04-20', fin='2026-05-29'),
    dict(p=3, unidad=3, trimestre=2, semanas=6, inicio='2026-06-01', fin='2026-07-10'),
    dict(p=4, unidad=4, trimestre=2, semanas=6, inicio='2026-08-04', fin='2026-09-11'),
    dict(p=5, unidad=5, trimestre=3, semanas=5, inicio='2026-09-14', fin='2026-10-23'),
    dict(p=6, unidad=6, trimestre=3, semanas=6, inicio='2026-10-26', fin='2026-12-04'),
]

# Nombres de columna tal cual estan en la fila de cabecera, normalizados.
AREAS = {
    'tutoria': 'tutoria', 'math': 'math', 'english': 'english',
    'social': 'social', 'comunicacion': 'comunicacion', 'science': 'science',
    'art': 'art', 'music': 'music', 'drama': 'drama', 'pe': 'pe',
    'spiritual history': 'spiritual', 'sh': 'spiritual',
}

def norm(s):
    s = unicodedata.normalize('NFKD', str(s))
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return re.sub(r'\s+', ' ', s).strip().lower()

def limpia(v):
    """Texto de celda a lista de lineas utiles."""
    if v is None:
        return []
    out = []
    for tramo in str(v).replace('\r', '\n').split('\n'):
        t = tramo.strip().strip('|').strip()
        t = re.sub(r'^[-*\u2022]\s*', '', t)
        t = re.sub(r'^\d+[.)]\s*', '', t)
        t = t.strip()
        if t and t not in ('-', '?'):
            out.append(t)
    return out

def cabecera(ws, fila=2):
    """Fila de cabecera: devuelve {columna -> area}. La columna sin titulo
    hereda la anterior (2.o tiene una segunda columna de Math sin cabecera)."""
    cols, ultima = {}, None
    for c in ws[fila]:
        t = norm(c.value or '')
        if t in AREAS:
            cols[c.column] = AREAS[t]
            ultima = AREAS[t]
        elif c.value is None and ultima and c.column > 4:
            cols[c.column] = ultima
    return cols

def hoja(ws, cab=2, grado_fijo=None):
    cols = cabecera(ws, cab)
    if not cols:
        return None
    grado = re.search(r'\d+', ws.title)
    filas, unidad = [], None
    for fila in ws.iter_rows(min_row=cab + 1):
        celdas = {c.column: c.value for c in fila}
        # columna B: marca de unidad (U1..U6)
        u = re.search(r'U(\d)', str(celdas.get(2) or ''))
        if u:
            unidad = int(u.group(1))
        # columna C: numero de semana dentro de la unidad
        try:
            semana = int(float(celdas.get(3)))
        except (TypeError, ValueError):
            continue
        if not unidad:
            continue
        contenido = {}
        for col, area in cols.items():
            for linea in limpia(celdas.get(col)):
                contenido.setdefault(area, [])
                if linea not in contenido[area]:
                    contenido[area].append(linea)
        filas.append(dict(unidad=unidad, semana=semana, areas=contenido))
    if grado_fijo is not None:
        return grado_fijo, filas
    return int(grado.group()) if grado else None, filas

def hojas(argv):
    """(hoja, fila de cabecera, grado forzado) de cada Excel de entrada."""
    for ws in openpyxl.load_workbook(Path(argv[0]), data_only=True).worksheets:
        yield ws, 2, None
    if len(argv) > 1 and argv[1].lower().endswith('.xlsx'):
        wb1 = openpyxl.load_workbook(Path(argv[1]), data_only=True)
        if 'Annual Plan' in wb1.sheetnames:
            # la hoja se llama "Annual Plan" a secas y su cabecera va en la
            # fila 1, no en la 2 como las de primaria
            yield wb1['Annual Plan'], 1, 1

def main():
    argv = [a for a in sys.argv[1:] if a.lower().endswith('.xlsx')]
    salidas = [a for a in sys.argv[1:] if a.lower().endswith('.json')]
    destino = Path(salidas[0]) if salidas else Path('scope/annual-plan-primary-2026.json')
    grados = {}
    for ws, cab, gfijo in hojas(argv):
        r = hoja(ws, cab, gfijo)
        if not r or not r[0]:
            continue
        grado, filas = r
        periodos = []
        for per in PERIODOS:
            semanas = []
            for n in range(1, per['semanas'] + 1):
                f = next((x for x in filas if x['unidad'] == per['unidad'] and x['semana'] == n), None)
                semanas.append(dict(n=n, areas=(f or {}).get('areas', {})))
            # lo que el Excel puso mas alla de las semanas reales del calendario
            sobra = [x for x in filas
                     if x['unidad'] == per['unidad'] and x['semana'] > per['semanas'] and x['areas']]
            periodos.append(dict(periodo=per['p'], unidad=per['unidad'],
                                 trimestre=per['trimestre'], inicio=per['inicio'],
                                 fin=per['fin'], semanas=semanas,
                                 fuera_de_calendario=[x['areas'] for x in sobra]))
        grados['g%d' % grado] = dict(grado=grado, periodos=periodos)
    salida = dict(
        fuente='Annual Plan Primary 2026.xlsx + EY_Assessment Criteria_G1.xlsx (hoja Annual Plan)'
               ' + Primary Calendar 2026.xlsx (Drive del colegio)',
        generado_por='tools/extrae_annual_plan.py',
        calendario=PERIODOS,
        grados=grados,
    )
    destino.parent.mkdir(parents=True, exist_ok=True)
    destino.write_text(json.dumps(salida, ensure_ascii=False, indent=1), encoding='utf-8')
    for k, v in sorted(grados.items(), key=lambda kv: kv[1]['grado']):
        con = sum(1 for p in v['periodos'] for s in p['semanas'] if s['areas'])
        tot = sum(len(p['semanas']) for p in v['periodos'])
        areas = sorted({a for p in v['periodos'] for s in p['semanas'] for a in s['areas']})
        print('%-4s %2d/%2d semanas con contenido · areas: %s' % (k, con, tot, ', '.join(areas)))
    print('->', destino, destino.stat().st_size // 1024, 'KB')

if __name__ == '__main__':
    main()
